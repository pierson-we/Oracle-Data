import openpyxl

def extract(file_name='/Users/wep/code/CTSI.xlsx', GOTB_column=5, enc_date_column=3):
    filepath = file_name  # create filepath from given directory, filename and extension
    wb = openpyxl.load_workbook(filepath)  # load the excel file
    data = {}  # 'data' dictionary holds all data extracted from file
    for sheet in wb.worksheets:  # loop through sheets in workbook
        sheet_dict = {}  # 'sheet_dict' holds all data extracted from the sheet
        for r in range(2, sheet.max_row + 1):  # loop through all rows in sheet. Each corresponds to single GOTB
            GOTB = get_GOTB(r, GOTB_column, wb, sheet)
            # need to look at data structure of REDCap export/import
            if GOTB not in sheet_dict.keys():
                sheet_dict[GOTB] = {}
            # need to look at data structure of REDCap export/import
            enc_date = sheet.cell(row=r, column=enc_date_column).value
            if enc_date not in sheet_dict[GOTB].keys():
                sheet_dict[GOTB][enc_date] = {}
            sample = len(sheet_dict[GOTB][enc_date].keys()) + 1 #mark the sample number for this encounter
            sample_dict = {}
            for c in range(1, sheet.max_column + 1):
                if c == GOTB_column: pass #ignore the GOTBID column
                cell_var_raw = sheet.cell(row=1, column=c).value #retrieve header for column
                cell_value_raw = sheet.cell(row=r, column=c).value #retrieve cell data
                #if cell_value_raw is None: pass #don't add to dict if cell is blank (don't think this is necessary)
                cell_var_REDCap = make_unicode(cell_var_raw)  # cell_var_REDCap = format_var(cell_var_raw, wb, sheet)
                if cell_var_REDCap == 'None': pass #if oracle variable doesn't correspond to REDCap variable, skip
                cell_value_REDCap = make_unicode(cell_value_raw)  # cell_value_REDCap = format_value(cell_value_raw, cell_var_REDCap, wb, sheet)
                sample_dict[cell_var_REDCap] = cell_value_REDCap #store sample info in dict
            sheet_dict[GOTB][enc_date][sample] = sample_dict
        data[sheet] = sheet_dict

    # return data
#    count = 1
#    for sheet in data.keys():
#        print sheet.title
#        for GOTB in sorted(data[sheet].iterkeys()):
#            for sample in data[sheet][GOTB]:
#                print "(%s) %s[%s]: %s" % (count, GOTB, sample, data[sheet][GOTB][sample])
#                count = count + 1
    return data


def make_unicode(cell_data):
    if type(cell_data) is str: #only format if value is str
        cell_data = cell_data.encode('utf-8')
        return cell_data
    else:
        return cell_data


def get_GOTB(r, GOTB_column, wb, sheet):
    return sheet.cell(row=r, column=GOTB_column).value


def format_var(cell_var_raw, wb, sheet):
    wb = openpyxl.load_workbook('variable_key_spreadsheet.xslx')
    sheet = wb.get_sheet_by_name('Sheet1')
    for c in range(1, sheet.max_column):
        if cell_var_raw == sheet.cell(row=1, column=c).value:
            return sheet.cell(row=2, column=c).value
        else:
            pass


def format_value(cell_value_raw, cell_var_REDCap, wb, sheet):
    if cell_var_REDCap in ['list of number variables']: return format_num(cell_value_raw)
    elif cell_var_REDCap in ['list of string variables']: return format_str(cell_value_raw)
    elif cell_var_REDCap in ['list of date variables']: return format_date(cell_value_raw)
    elif cell_var_REDCap in ['list of time variables']: return format_time(cell_value_raw)
    else:
        pass


def format_num(cell_value_raw):
    return 'stripped of all blank space and non-numeric characters'


def format_str(cell_value_raw):
    return 'probably actually dont need to do anything here'


def format_date(cell_value_raw):
    return 'consistent mm/dd/yyyy format'


def format_time(cell_value_raw):
    return 'python instant'


def store(data_in, column_format_xlsx, save_filepath):
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    dest_filename = save_filepath

    column_wb = openpyxl.load_workbook(column_format_xlsx)
    column_sheet = column_wb.get_sheet_by_name('Sheet1')
    column_dict = {}

    for c in range(1, column_sheet.max_column + 1):
        column_dict[make_unicode(column_sheet.cell(row = 1, column = c).value)] = c

    sheet_count = 0
    for sheet in data_in:

        # this section is cosmetic-handle naming of sheets in case of multiple sheets in source file
        sheet_count = sheet_count + 1
        if sheet_count == 1:
            ws = wb.active
        else: ws = wb.create_sheet(title='Sheet %s' % (sheet_count))

        #store column headers
        for header in column_dict.keys():
            ws.cell(row = 1, column = column_dict[header]).value = header

        row = 1
        GOTB_count = 0
        for GOTB in sorted(data_in[sheet].iterkeys()):
            GOTB_count = GOTB_count + 1
            for encounter in data_in[sheet][GOTB]:
                for sample in data_in[sheet][GOTB][encounter]:
                    row = row + 1
                    for var in sorted(data_in[sheet][GOTB][encounter][sample].iterkeys()):
                        value = data_in[sheet][GOTB][encounter][sample][var]
                        ws.cell(row=row, column=column_dict[var]).value = value
        print 'Patient Count: %s\nSample Count: %s' % (GOTB_count, row)


    wb.save(filename = dest_filename)
    return True
